# -*- coding: utf-8 -*-

"""
@author: shencheng
@software: PyCharm
@description: 操作、读取excel的工具类
@time: 2023/12/07 9:34
"""
from common import dir_util
import openpyxl


class ExcelUtil:

    def excel_read(self, filename):
        filepahth=dir_util.testdata_dir + filename + '.xlsx'
        print(filepahth)
        wb = openpyxl.load_workbook(dir_util.testdata_dir + filename + '.xlsx')
        sheet = wb.worksheets[0]
        row = sheet.max_row
        col = sheet.max_column
        row_list = []
        for r in range(2, row + 1):
            col_list = []
            for col in range(1, col + 1):
                datavalue = sheet.cell(row=r, column=col).value
                if datavalue is None:
                    datavalue = ''
                    col_list.append(datavalue)
                else:
                    col_list.append(datavalue)
            row_list.append(col_list)
        return row_list

if __name__ == '__main__':
    res=ExcelUtil().excel_read('pm_phone')
    print(res)

